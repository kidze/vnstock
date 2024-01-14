import pandas as pd
import numpy as np
from vnstock import *
import json
from openpyxl import load_workbook

# Read the CSV file
df = pd.read_csv("vn_stock_companies.csv")

# Filter the rows where group_code is 'VNINDEX' or 'HNX'
vnindex_df = df[df["group_code"] != "UpcomIndex"]

# Get the list of tickers
symbollist = vnindex_df["ticker"].tolist()

# Filter the symbol list
filtered_symbollist = [symbol for symbol in symbollist if len(symbol) <= 4]

financial_data = []  # List to store the financial data objects

# Limit the number of stocks
filtered_symbollist = ['VIC','KSV','VPG','DGC','DHC','DGW','DGC','HPG','CTR','VCS','CAP','RAL','DHA','FPT']

for symbol in filtered_symbollist:
    net_income = {}
    compound_rate = {}
    industry = ""
    dividend_yield = 0.0  # Default value for dividend_yield
    try:
        overview = company_overview(symbol)
        industry = overview.loc[0, "industry"]
    except Exception as e:
        print(f"Error fetching data for symbol {symbol} in company_overview: {str(e)}")

    try:
        # data = financial_report(
        #     symbol=symbol, report_type="IncomeStatement", frequency="Yearly"
        # )

        # # Remove the 'Q5 ' from years
        # data.columns = [col.replace("Q5 ", "") for col in data.columns]

        # # Find the row index where "CHỈ TIÊU" matches one of the three values
        # row_index = data[
        #     data["CHỈ TIÊU"].str.contains(
        #         "Lợi nhuận của Cổ đông của Công ty mẹ|Lợi nhuận sau thuế của chủ sở hữu, tập đoàn|Lợi nhuận sau thuế phân bổ cho chủ sở hữu|Lợi nhuận sau thuế"
        #     )
        # ].index[0]

        # # Use the row index to get the net income data
        # net_income = data.loc[row_index].to_dict()

        # # Remove the first entry in the dictionary
        # del net_income["CHỈ TIÊU"]
        # years = list(net_income.keys())[5:]

        # compound_rate = {}

        # for year in years:
        #     year_int = int(year)
        #     sum_net_income_5_years = sum(
        #         net_income[str(y)] for y in range(year_int - 5, year_int)
        #     )
        #     compound_rate[year] = (
        #         net_income[year] - net_income[str(year_int - 5)]
        #     ) / sum_net_income_5_years
            
        data = financial_flow(symbol=symbol, report_type="incomestatement", report_range="yearly")
        data.index = data.index.str.replace("-Q5", "")
        
        net_income_column = 'postTaxProfit'  # Update with the actual column name
                
        for year in data.index:
            year_int = int(year)
            net_income[year] = data.loc[year, net_income_column]
            if str(year_int - 5) in data.index:
                sum_net_income_5_years = sum(data.loc[str(y), net_income_column] for y in range(year_int - 5, year_int))
                if sum_net_income_5_years != 0:
                    compound_rate[year] = (data.loc[year, net_income_column] - data.loc[str(year_int - 5), net_income_column]) / sum_net_income_5_years
                else:
                    compound_rate[year] = 0  # Set compound rate to zero or another value if appropriate

        
    except Exception as e:
        print(f"Error fetching data for symbol {symbol} in financial_report: {str(e)}")

    final_average_roe = 0.0
    final_average_roc = 0.0
    pe = None
    debtOnCapital = None
    try:
        # Calculate average 5 year ROE
        df = financial_ratio(symbol, "yearly", True)

        if "roe" in df.columns and len(df) >= 5:
            average_roe = df["roe"].head(5).mean()
            ema_roe = df["roe"].head(5).ewm(span=3, adjust=False).mean()
            if not np.isnan(ema_roe.iloc[-1]):
                final_average_roe = ema_roe.iloc[-1]
        else:
            average_roe = df["roe"].mean()
            ema_roe = df["roe"].ewm(span=len(df), adjust=False).mean()
            if not np.isnan(ema_roe.iloc[-1]):
                final_average_roe = ema_roe.iloc[-1]

        # Get latest Price to Earning ratio and dividend yield and Debt On Capital from Debt on Equity
        # pe ratio is taken from financial_ratio quarterly
        df_quarterly = financial_ratio(symbol, "quarterly", False)
        pe = df_quarterly.loc[0, "priceToEarning"]

        dividend_yield = df.loc[0, "dividend"]

        debtOnEquity = df.loc[0, "debtOnEquity"]
        if debtOnEquity is not None:
            debtOnCapital = debtOnEquity / (debtOnEquity + 1)
            final_average_roc = final_average_roe * (1 - debtOnCapital)

    except Exception as e:
        print(f"Error fetching data for symbol {symbol} in financial_ratio: {str(e)}")

    financial_data.append(
        {
            "ticker": symbol,
            "industry": industry,
            "net_income": net_income,
            "compound_rate": compound_rate,
            "average_5y_roc": final_average_roc,
            "average_5y_roe": final_average_roe,
            "pe": pe,
            "dividend_yield": dividend_yield,
            "debtOnCapital": debtOnCapital,
        }
    )

# Process financial_data to calculate average_5y_compound_rate
for entry in financial_data:
    compound_rates = entry["compound_rate"]
    available_years = list(compound_rates.keys())
    if len(available_years) >= 5:
        latest_years = list(compound_rates.keys())[:5]
        average_5y_compound_rate = sum(compound_rates[year] for year in latest_years) / 5
    else:
        available_rates = [compound_rates[year] for year in available_years]
        if len(available_rates) > 0:
            average_5y_compound_rate = sum(available_rates) / len(available_rates)
        else:
            average_5y_compound_rate = 0  # Set a default value when there's no data
    
    entry["average_5y_compound_rate"] = average_5y_compound_rate

# sort financial_data based on the highest roc and average_5y_compound_rate
financial_data = sorted(
    financial_data,
    key=lambda x: (x["average_5y_roc"], x["average_5y_compound_rate"]),
    reverse=True,
)

# Define the file path to save the data
file_path = "financial_data.json"

# Write financial_data to the file in JSON format
with open(file_path, "w") as file:
    json.dump(financial_data, file, default=str)

# Assuming financial_data is a dictionary
df = pd.DataFrame(financial_data)

# Write the DataFrame to a CSV file
df.to_csv("financial_data.csv", index=False)


# ########################## Replacing the financial_data sheet in the excel. Removed for now.
# # Load the existing workbook
# book = load_workbook('stock-valuation-2023.xlsx')
# # Create an Excel writer object with the loaded workbook
# writer = pd.ExcelWriter('stock-valuation-2023.xlsx', engine='openpyxl')
# writer.book = book

# # Delete the 'financial_data' sheet if it exists
# if 'financial_data' in book.sheetnames:
#     std = book['financial_data']
#     book.remove(std)

# # Write the DataFrame to the 'financial_data' sheet
# df.to_excel(writer, sheet_name='financial_data', index=False)

# # Save the workbook
# writer.save()
